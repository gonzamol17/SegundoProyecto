import unittest
from pathlib import Path

import openpyxl
import xlrd
import time


class MyTestCase(unittest.TestCase):
    def test_something(self):
        # Esta parte toma todas las filas y columnas del archivo excel
        userExcel_directory = Path(__file__).parent.parent
        excelUser_file_path = userExcel_directory.parent.parent / "Datos" / "Users.xlsx"
        inputWorkbook = openpyxl.load_workbook(excelUser_file_path)
        inputWorksheet = inputWorkbook.active

        #row = inputWorksheet.rows
        #col = inputWorksheet.cols

        # Obtener número de filas y columnas
        rows = inputWorksheet.max_row
        cols = inputWorksheet.max_column

        print("\n")
        print("números de filas: "+str(rows))
        print("números de columnas: "+str(cols))

        print(inputWorksheet.cell(row=2, column=1).value)  # Para obtener el valor de la celda A2
        print(inputWorksheet.cell(row=2, column=2).value)  # Para obtener el valor de la celda B2

        # Imprimir todos los valores de las celdas en la segunda fila
        for col in range(1, cols + 1):  # Iterar sobre las columnas
            print(inputWorksheet.cell(row=2, column=col).value)  # Imprimir valores de la segunda fila
