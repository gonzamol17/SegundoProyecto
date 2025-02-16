import unittest
from pathlib import Path

from selenium import webdriver
import unittest
import sys
import os

from selenium.webdriver.chrome.service import Service

sys.path.append(os.path.join(os.path.dirname(__file__), "../..", ".."))
import openpyxl
import time
from POM.Pages.LandingPage import LandingPage
from POM.Pages.LoginPage import LoginPage
from POM.Pages.MyAccountPage import MyAccountPage


class LoginWithExternalFile(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
<<<<<<< HEAD
        #Esto quedó deprecado y utilicé la opción que está abajo
        # cls.driver = webdriver.Chrome("/Drivers/chromedriver.exe")
        # cls.driver.implicitly_wait(10)
        # cls.driver.maximize_window()
        # service = Service(executable_path="C:\\Users\\User\\PycharmProjects\\SegundoProyecto\\Drivers\\chromedriver.exe")

        current_directory = Path(__file__).parent.parent
        # Construir la ruta relativa al 'chromedriver.exe' desde el directorio del script
        driver_path = current_directory.parent.parent / "Drivers" / "chromedriver.exe"
        # Inicializar el servicio de Chrome con la ruta construida
        service = Service(executable_path=str(driver_path))
        cls.driver = webdriver.Chrome(service=service)
=======
        cls.driver = webdriver.Chrome()
        #cls.driver = webdriver.Chrome("/Drivers/chromedriver.exe")
>>>>>>> 2fa78c339105440924a5bb0d77d85a97438def7b
        cls.driver.implicitly_wait(10)
        cls.driver.maximize_window()

    def test_LoginUsingXmlFile(self):
        driver = self.driver
        driver.get("https://automationteststore.com/")
        time.sleep(2)
        lp = LandingPage(driver)
        lp.click_Go_Login()
        logpa = LoginPage(driver)
        time.sleep(2)
        my = MyAccountPage(driver)

        #Esta parte toma todas las filas y columnas del archivo excel
<<<<<<< HEAD
        #path = "C:\\Users\\User\\PycharmProjects\\SegundoProyecto\\Datos\\Users.xlsx"

        # Estas dos lineas reemplazan la anterior para obtener el path relativo del archivo excel
        userExcel_directory = Path(__file__).parent.parent
        excelUser_file_path = userExcel_directory.parent.parent / "Datos" / "Users.xlsx"
=======
        path = "C:\\Users\\GonzaloJavierMolinaC\\OneDrive - Capitole Consulting\\Escritorio\\Automation Practice\\pythonprojects\\SegundoProyecto\\Datos\\Users.xlsx"
        inputWorkbook = xlrd.open_workbook(path)
        inputWorksheet = inputWorkbook.sheet_by_index(0)
        row = inputWorksheet.nrows
        col = inputWorksheet.ncols
>>>>>>> 2fa78c339105440924a5bb0d77d85a97438def7b



        #inputWorkbook = openpyxl.load_workbook(path)
        # inputWorkbook = xlrd.open_workbook(path)
        # inputWorksheet = inputWorkbook.sheet_by_index(0)
        # #sheet1= inputWorksheet.sheet_by_index(0)
        # row = inputWorksheet.nrows
        # col = inputWorksheet.ncols
        #
        #
        # for aux in range(1, inputWorksheet.nrows):
        #     print(aux)
        #     username = inputWorksheet.cell_value(aux, 1)
        #     password = inputWorksheet.cell_value(aux, 2)
        #     time.sleep(4)
        #     logpa.do_Login(username, password)
        #     time.sleep(4)
        #     my.seleccionar_Logoff()
        #     time.sleep(2)
        #     lp.click_Go_Login()
        #     time.sleep(2)

        # Abre el archivo .xlsx usando openpyxl
        inputWorkbook = openpyxl.load_workbook(excelUser_file_path)
        inputWorksheet = inputWorkbook.active

        # Itera a través de las filas, comenzando desde la segunda fila (para saltarte los encabezados)
        for aux in range(2, inputWorksheet.max_row + 1):
            username = inputWorksheet.cell(row=aux, column=2).value
            password = inputWorksheet.cell(row=aux, column=3).value
            time.sleep(2)
            logpa.do_Login(username, password)
            time.sleep(4)
            my.seleccionar_Logoff()
            time.sleep(2)
            lp.click_Go_Login()
            time.sleep(2)

    @classmethod
    def tearDownClass(cls):
        cls.driver.close()
        cls.driver.quit()
        print("Test Completed")
<<<<<<< HEAD
=======


# if __name__ == '__main__':
#      unittest.main(testRunner=HtmlTestRunner.HTMLTestRunner(output='C:\\Users\\admin\\PycharmProjects\\SegundoProyecto\\Reports'), verbosity=2)
#
>>>>>>> 2fa78c339105440924a5bb0d77d85a97438def7b
